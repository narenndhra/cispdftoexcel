#!/usr/bin/env python3
"""
CIS Benchmark PDF to Excel Converter
=====================================
Universal converter for ANY CIS Benchmark PDF (Windows, Linux, Cloud, Network devices, etc.)

Features:
- Auto-detects CIS benchmark sections
- Extracts all components: Description, Rationale, Impact, Audit (CLI & GUI), Remediation, Default Value, References
- Creates multi-tab Excel workbook (one tab per section)
- Professional formatting with color coding
- Handles variations in CIS PDF formats

Usage:
    python cis_pdf_to_excel_converter.py <input_pdf> [output_excel]

Example:
    python cis_pdf_to_excel_converter.py CIS_Ubuntu_22_04.pdf
    python cis_pdf_to_excel_converter.py CIS_Windows_Server_2022.pdf CIS_Win_2022_Audit.xlsx

Author: Security Audit Team
Version: 1.0
"""

import sys
import re
import json
import argparse
from pathlib import Path
from datetime import datetime
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CISBenchmarkExtractor:
    """Extract recommendations from CIS Benchmark PDFs"""
    
    def __init__(self, pdf_path):
        self.pdf_path = Path(pdf_path)
        self.reader = PdfReader(str(pdf_path))
        self.recommendations = []
        self.benchmark_title = ""
        self.benchmark_version = ""
        
    def extract_metadata(self):
        """Extract benchmark title and version from PDF"""
        # Check first 5 pages for title
        for page_num in range(min(5, len(self.reader.pages))):
            text = self.reader.pages[page_num].extract_text()
            
            # Look for CIS benchmark title pattern
            title_match = re.search(r'CIS\s+(.+?)\s+Benchmark', text, re.IGNORECASE)
            if title_match:
                self.benchmark_title = f"CIS {title_match.group(1)} Benchmark"
            
            # Look for version
            version_match = re.search(r'v(\d+\.\d+\.\d+)', text)
            if version_match:
                self.benchmark_version = f"v{version_match.group(1)}"
        
        print(f"📄 Detected: {self.benchmark_title} {self.benchmark_version}")
    
    def find_recommendations_start_page(self):
        """Find where recommendations section starts"""
        for page_num in range(min(30, len(self.reader.pages))):
            text = self.reader.pages[page_num].extract_text()
            
            # Look for first recommendation pattern (1.1, 1.1.1, etc.)
            if re.search(r'\n1\.1(?:\.\d+)?\s+[A-Z]', text):
                print(f"📍 Recommendations start at page {page_num + 1}")
                return page_num
        
        return 15  # Default fallback
    
    def extract_recommendations(self):
        """Extract all recommendations from PDF"""
        start_page = self.find_recommendations_start_page()
        
        print(f"\n🔍 Extracting recommendations from {len(self.reader.pages)} pages...")
        
        for page_num in range(start_page, len(self.reader.pages)):
            text = self.reader.pages[page_num].extract_text()
            
            # Pattern for recommendation headers
            # Matches: "1.1 Title (Automated)" or "2.3.4 Title (Manual)" or "1.1 Title (Scored)" etc.
            pattern = r'(?:^|\n)(\d+(?:\.\d+)+)\s+([A-Z][^\n]+?)\s*\((Automated|Manual|Scored|Not Scored)\)'
            
            matches = re.finditer(pattern, text)
            
            for match in matches:
                num = match.group(1)
                title = match.group(2).strip()
                status = match.group(3)
                
                # Extract content after header (next 3000 chars)
                start_pos = match.end()
                content = text[start_pos:start_pos+3500]
                
                # Extract all components
                rec = self._extract_recommendation_details(num, title, status, content)
                
                if rec and rec['audit']:  # Only add if has audit steps
                    self.recommendations.append(rec)
        
        # Remove duplicates
        self._remove_duplicates()
        
        # Sort by number
        self.recommendations.sort(key=lambda x: [int(p) for p in x['num'].split('.')])
        
        print(f"✅ Extracted {len(self.recommendations)} complete recommendations")
    
    def _extract_recommendation_details(self, num, title, status, content):
        """Extract all details for a single recommendation"""
        
        # Extract Profile/Level
        profile_match = re.search(r'•?\s*(Level \d+|Profile Applicability)', content)
        profile = profile_match.group(1) if profile_match else "Level 1"
        
        # Extract Description
        desc_match = re.search(r'Description:\s*(.+?)(?:\nRationale:|\nImpact:|\nAudit:)', content, re.DOTALL)
        description = desc_match.group(1).strip() if desc_match else ""
        
        # Extract Rationale
        rat_match = re.search(r'Rationale:\s*(.+?)(?:\nImpact:|\nAudit:)', content, re.DOTALL)
        rationale = rat_match.group(1).strip() if rat_match else ""
        
        # Extract Impact
        impact_match = re.search(r'Impact:\s*(.+?)(?:\nAudit:|\nRemediation:)', content, re.DOTALL)
        impact = impact_match.group(1).strip() if impact_match else ""
        
        # Extract Audit
        audit_match = re.search(r'Audit:\s*(.+?)(?:\nRemediation:|\nDefault Value:)', content, re.DOTALL)
        audit = audit_match.group(1).strip() if audit_match else ""
        
        # Extract Remediation
        remed_match = re.search(r'Remediation:\s*(.+?)(?:\nDefault Value:|\nImpact:|\nReferences:)', content, re.DOTALL)
        remediation = remed_match.group(1).strip() if remed_match else ""
        
        # Extract Default Value
        default_match = re.search(r'Default Value:\s*(.+?)(?:\nReferences:|\nCIS Controls:)', content, re.DOTALL)
        default_value = default_match.group(1).strip() if default_match else ""
        
        # Extract References
        ref_match = re.search(r'References:\s*(.+?)(?:\nCIS Controls:|\nAdditional Information:)', content, re.DOTALL)
        references = ref_match.group(1).strip() if ref_match else ""
        
        return {
            'num': num,
            'title': title,
            'profile': profile,
            'status': status,
            'description': description[:1500],
            'rationale': rationale[:1000],
            'impact': impact[:1000],
            'audit': audit[:2500],
            'remediation': remediation[:1500],
            'default_value': default_value[:500],
            'references': references[:800]
        }
    
    def _remove_duplicates(self):
        """Remove duplicate recommendations"""
        seen = set()
        unique = []
        for rec in self.recommendations:
            if rec['num'] not in seen:
                seen.add(rec['num'])
                unique.append(rec)
        self.recommendations = unique
    
    def organize_by_section(self):
        """Organize recommendations by main section"""
        sections = {}
        
        for rec in self.recommendations:
            # Get main section (e.g., "1" from "1.2.3")
            main_section = rec['num'].split('.')[0]
            
            if main_section not in sections:
                sections[main_section] = []
            
            sections[main_section].append(rec)
        
        return sections


class ExcelWorkbookGenerator:
    """Generate Excel workbook from CIS recommendations"""
    
    # Section name templates (can be customized)
    SECTION_NAMES = {
        '1': 'Initial Setup',
        '2': 'System Configuration',
        '3': 'Network & Services',
        '4': 'Security Profiles',
        '5': 'Access Control',
        '6': 'Authentication',
        '7': 'Logging & Monitoring',
        '8': 'System Maintenance',
        '9': 'Additional Hardening'
    }
    
    def __init__(self, recommendations, benchmark_title, benchmark_version):
        self.recommendations = recommendations
        self.benchmark_title = benchmark_title
        self.benchmark_version = benchmark_version
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
    
    def create_index_sheet(self, sections):
        """Create index/overview sheet"""
        ws = self.wb.create_sheet('📑 INDEX')
        
        # Title
        ws.merge_cells('A1:E1')
        title = ws.cell(1, 1, f'{self.benchmark_title} {self.benchmark_version}')
        title.font = Font(size=16, bold=True, color='FFFFFF')
        title.fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Subtitle
        ws.merge_cells('A2:E2')
        subtitle = ws.cell(2, 1, f'Audit Checklist - Generated on {datetime.now().strftime("%Y-%m-%d")}')
        subtitle.font = Font(size=11, italic=True)
        subtitle.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 25
        
        ws.append([])
        
        # Headers
        headers = ['Section', 'Section Name', 'Controls', 'Tab Name', 'Description']
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(4, col)
            cell.font = Font(size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 32
        ws.column_dimensions['E'].width = 50
        
        # Add section rows
        row = 5
        for sec_num in sorted(sections.keys(), key=int):
            sec_name = self.SECTION_NAMES.get(sec_num, f'Section {sec_num}')
            sec_recs = sections[sec_num]
            sheet_name = f"{sec_num}. {sec_name}"[:31]
            
            ws.append([sec_num, sec_name, len(sec_recs), sheet_name, ''])
            
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.font = Font(size=10)
                cell.alignment = Alignment(vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            ws.row_dimensions[row].height = 30
            row += 1
    
    def create_section_sheet(self, section_num, section_name, recommendations):
        """Create sheet for a section"""
        sheet_name = f"{section_num}. {section_name}"[:31]
        ws = self.wb.create_sheet(sheet_name)
        
        # Title
        ws.merge_cells('A1:H1')
        title = f'{self.benchmark_title} - Section {section_num}: {section_name}'
        cell = ws.cell(1, 1, title)
        cell.font = Font(size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='00518F', end_color='00518F', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Headers
        headers = ['#', 'Control Title', 'Level', 'Description & Impact', 
                  'Audit Steps (CLI & GUI)', 'Remediation', 'Default Value', 'References/Status']
        ws.append(headers)
        
        for col in range(1, 9):
            cell = ws.cell(2, col)
            cell.font = Font(size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        ws.row_dimensions[2].height = 40
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 55
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 40
        
        # Add recommendations
        row = 3
        for rec in recommendations:
            self._add_recommendation_row(ws, row, rec)
            row += 1
    
    def _add_recommendation_row(self, ws, row, rec):
        """Add a recommendation row"""
        # Combine description, rationale, impact
        desc_full = rec['description']
        if rec['rationale']:
            desc_full += f"\n\nRATIONALE:\n{rec['rationale']}"
        if rec['impact']:
            desc_full += f"\n\nIMPACT:\n{rec['impact']}"
        
        data = [
            rec['num'],
            rec['title'],
            rec['profile'],
            desc_full,
            rec['audit'],
            rec['remediation'],
            rec['default_value'],
            rec['references']
        ]
        
        ws.append(data)
        
        # Style cells
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Color code Level column
            if col == 3:
                if 'Level 1' in rec['profile']:
                    cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                    cell.font = Font(size=9, bold=True)
                elif 'Level 2' in rec['profile']:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.font = Font(size=9, bold=True)
        
        # Dynamic row height
        max_lines = max(
            len(desc_full.split('\n')),
            len(rec['audit'].split('\n')),
            len(rec['remediation'].split('\n'))
        )
        ws.row_dimensions[row].height = min(max(max_lines * 14, 80), 350)
    
    def generate(self, sections, output_path):
        """Generate complete workbook"""
        print(f"\n📊 Generating Excel workbook...")
        
        # Create index
        self.create_index_sheet(sections)
        print(f"   ✓ Created index sheet")
        
        # Create section sheets
        for sec_num in sorted(sections.keys(), key=int):
            sec_name = self.SECTION_NAMES.get(sec_num, f'Section {sec_num}')
            sec_recs = sections[sec_num]
            
            self.create_section_sheet(sec_num, sec_name, sec_recs)
            print(f"   ✓ Created: Section {sec_num} - {sec_name} ({len(sec_recs)} controls)")
        
        # Save
        self.wb.save(output_path)
        print(f"\n✅ Excel workbook saved: {output_path}")
        print(f"   Total Sheets: {len(sections) + 1} (1 Index + {len(sections)} Sections)")
        print(f"   Total Controls: {len(self.recommendations)}")


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description='Convert CIS Benchmark PDF to Excel Audit Checklist',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s CIS_Ubuntu_22_04.pdf
  %(prog)s CIS_Windows_Server_2022.pdf CIS_Win_2022_Audit.xlsx
  %(prog)s CIS_FortiGate_7_4.pdf --output FortiGate_Audit.xlsx

The script will:
  1. Extract all CIS recommendations from the PDF
  2. Organize them by section
  3. Create a multi-tab Excel workbook with complete audit details
  4. Include: Description, Impact, Audit steps (CLI & GUI), Remediation, References
        '''
    )
    
    parser.add_argument('input_pdf', help='Path to CIS Benchmark PDF file')
    parser.add_argument('output_excel', nargs='?', help='Output Excel file path (optional)')
    parser.add_argument('--output', '-o', help='Output Excel file path (alternative)')
    
    args = parser.parse_args()
    
    # Determine output path
    if args.output_excel:
        output_path = args.output_excel
    elif args.output:
        output_path = args.output
    else:
        # Auto-generate output filename
        input_path = Path(args.input_pdf)
        output_path = input_path.stem + '_Audit_Checklist.xlsx'
    
    print("="*80)
    print("CIS BENCHMARK PDF TO EXCEL CONVERTER")
    print("="*80)
    print(f"📁 Input PDF: {args.input_pdf}")
    print(f"💾 Output Excel: {output_path}")
    print()
    
    try:
        # Extract recommendations
        extractor = CISBenchmarkExtractor(args.input_pdf)
        extractor.extract_metadata()
        extractor.extract_recommendations()
        
        # Organize by section
        sections = extractor.organize_by_section()
        
        print(f"\n📂 Organized into {len(sections)} sections:")
        for sec in sorted(sections.keys(), key=int):
            print(f"   Section {sec}: {len(sections[sec])} controls")
        
        # Generate Excel workbook
        generator = ExcelWorkbookGenerator(
            extractor.recommendations,
            extractor.benchmark_title,
            extractor.benchmark_version
        )
        generator.generate(sections, output_path)
        
        print("\n" + "="*80)
        print("✅ CONVERSION COMPLETE!")
        print("="*80)
        print(f"\n📄 Your audit checklist is ready: {output_path}")
        print(f"   • Section-wise organization (one tab per section)")
        print(f"   • Complete audit details included")
        print(f"   • Professional formatting with color coding")
        print(f"   • Ready for audit use!")
        
    except FileNotFoundError:
        print(f"❌ ERROR: File not found: {args.input_pdf}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
